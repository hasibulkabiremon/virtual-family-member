import speech_recognition as sr
import pyttsx3
import pywhatkit
import datetime
import wikipedia
import pyjokes
import openpyxl
import pyaudio
import nltk
import warnings
from nltk.tokenize import word_tokenize
import nltk
# from newspaper import Article
import string
import random
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
# import PySimpleGUI as sg


# nltk.download('punkt', quiet=True)
# nltk.download('wordnet', quiet=True)

engine = pyttsx3.init()
friend = pyttsx3.init()
r = sr.Recognizer()
m = sr.Microphone()



voices = engine.getProperty('voices')
engine.setProperty('voice', voices[0].id)


# start searchbot

# create a function to return lower case words
def LemNormalize(text):
    return nltk.word_tokenize(text.lower().translate(remove_punct_dict))


def greeting(sentence):
    for word in sentence.split():
        if word.lower() in greeting_input:
            return random.choice(greeting_response)


def response(user_response):
    # user response and robo responce
    # user_response="WHat is chronic disease"
    # print(user_response)
    # robo response
    robo_response = ''
    sent_tokens.append(user_response)
    # print(sent_tokens)
    tfidfvec = TfidfVectorizer(tokenizer=LemNormalize, stop_words='english')
    tfidf = tfidfvec.fit_transform(sent_tokens)
    # print(tfidf)
    # get similarity score
    val = cosine_similarity(tfidf[-1], tfidf)
    # print(val)
    idx = val.argsort()[0][-2]
    flat = val.flatten()
    flat.sort()
    score = flat[-2]
    # print(score)
    if score == 0:
        print ("sorry,i dont understand")
        talk("sorry,i dont understand")
        engine.runAndWait()

    else:
        robo_response = robo_response + sent_tokens[idx]

    sent_tokens.remove(user_response)
    return robo_response




def greeting(sentence):
    for word in sentence.split():
        if word.lower() in greeting_input:
            return random.choice(greeting_response)


def finalresponse(command):
    user_response = command
    user_response = user_response.lower()

    warnings.filterwarnings('ignore')

    # download package from nltk
    nltk.download('punkt', quiet=True)
    nltk.download('wordnet', quiet=True)

    userdata = open('temporary.txt', 'r')
    converteduserdata = userdata.read()

    text = converteduserdata
    global sent_tokens
    sent_tokens = nltk.sent_tokenize(text)
    # print(sent_tokens)

    # creating a dictionary to remove the punctuation
    global remove_punct_dict
    remove_punct_dict = dict((ord(punct), None) for punct in string.punctuation)

    # user_response=user_response.lower()
    if (user_response != 'bye'):
        if (user_response == 'thanks' or user_response == 'thank you'):

            print(" anytime :)")
        else:
            if (greeting(user_response) != None):
                print(greeting(user_response))
            else:
                reply = response(user_response)
                reply = reply.replace('i ', target_name+' ')
                reply = reply.replace(' am ', ' is ')
                reply = reply.replace(' have ', ' has ')
                reply = reply.replace(' my ',' his ')
                reply = reply.replace(' is going to ',' has left for ')

                if user in reply:
                    reply = reply.replace(' '+user+' ',' your ')
                    reply = reply.replace(user + '\'s', 'your')
                    reply = reply.replace(' ' + user, 'your')
                    reply = reply.replace('she is', 'you are')
                    reply = reply.replace('he is', 'you are')

                print(reply)
                talk(reply)
                engine.runAndWait()
    else:
        flag = False
        print("FamBot: see you later :)")
    open('temporary.txt', 'w').close()


# end searchbot


def talk(text):
    engine.say(text)
    engine.runAndWait()


def take_command():
    try:
        command = input()
        command = command.lower()
        if 'alexa' in command:
            command = command.replace('alexa', '')
            command = command.lower()
            print(command)
    except:
        command = 'I could\'t hear properly'
    return command


def run_intelligent():
    time = datetime.datetime.now().strftime('%H:%M:%S')
    date = datetime.datetime.now().strftime('%Y-%m-%d')

    command = voice_input()
    command = command.lower()
    print(command)

    for member in entry:
        if member in command:
            outputsheet = wb['Sheet1']
            outputrow = outputsheet.max_row

            for i in range(1, outputrow + 1):
                name = str(outputsheet.cell(i, 1).value)

                if member == name:
                    global target_name
                    target_name=member
                    speech = str(outputsheet.cell(i, 2).value)
                    with open('temporary.txt', 'a+') as file:
                        print(speech + '.', file=file)
            finalresponse(command)

    if 'play' in command:
        song = command.replace('play', '')
        talk('playing ' + song)
        pywhatkit.playonyt(song)
    elif command=="thank you":
        talk("Welcome")
        print("Welcome")
        exit()

    elif 'time' in command:
        time = datetime.datetime.now().strftime('%I:%M %p')
        talk('Current time is ' + time)
    elif 'who' in command:
        person = command.replace('who the heck is', '')
        info = wikipedia.summary(person, 1)
        print(info)
        talk(info)
    elif 'date' in command:
        talk('sorry, I have a headache')
    elif 'are you single' in command:
        talk('I am in a relationship with wifi')
    elif 'joke' in command:
        talk(pyjokes.get_joke())
    elif 'remember' in command:
        command = command.replace('remember', '')
        inputsheet = wb['Sheet1']
        inputrow = inputsheet.max_row

        inputsheet.cell(row=inputrow + 1, column=1, value=user)
        inputsheet.cell(row=inputrow + 1, column=2, value=command)
        inputsheet.cell(row=inputrow + 1, column=3, value=date)
        inputsheet.cell(row=inputrow + 1, column=4, value=time)

        wb.save("data.xlsx")
        talk('Ok')

        # with open(user + '.txt', 'a+') as file:
        #     print(command + '.', file=file)

    elif 'bye' in command:
        exit()
    else:
        talk('Please say again.')


def voice_input():
    try:
        with sr.Microphone() as source:
            print('Listening...')
            listener.pause_threshold = 1
            listener.energy_threshold = 3000
            audio = listener.listen(source)
            # print(audio)
        try:
            text = listener.recognize_google(audio, language='en-in')

            # with open(user + '.txt', 'a') as file:
            #     print((text + '.'), file=file)

            print(f'You said :{text}')
            # friend.say(text)
            # friend.runAndWait()
        except:
            print('Sorry I couldn\'t recognize your voice')
        return text
    except:
        return 'ok'


# workbook
wb = openpyxl.load_workbook('data.xlsx')
s1 = wb['Sheet1']

# Member Record
entry = []
row = s1.max_row
#
for i in range(1, row + 1):
    x = str(s1.cell(i, 1).value)
    with open('temporary.txt', 'a+') as file:
        entry.append(x)
entry = list(dict.fromkeys(entry))
#print(entry)

# keywords for greetings
greeting_input = ["hi", "hello", "hey", "hola"]
greeting_response = ["howdy", "hey there", "hi", "hello :)"]

# Layout

talk('hi, What\'s your name?')
print('hi, What\'s your name?')


while True:
    listener = sr.Recognizer()
    # talk('hi, What\'s your name?')
    user = voice_input()
    user = user.lower()
    user_token = word_tokenize(user)
    # print(user_token)
    user = user_token[-1]

    print('You are ' + user + '. Am I right?')
    talk('You are ' + user + '. Am I right?')

    comment = voice_input()
    print(comment)

    if 'yes' in comment:
        print('yes')
        while True:
            talk("Hello"+user+"How can i help You?")
            print("Hello"+user+"How can i help You?")
            run_intelligent()
    else:
        talk("Sorry I Don't understand. Please tell me your name again")
        print("Sorry I Don't understand. Please tell me your name again")



        # with sr.Microphone() as source:
        #     print('Listening...')
        #     listener.pause_threshold = 1
        #     listener.energy_threshold = 4000
        #     audio = listener.listen(source)
        #     # print(audio)
        # try:
        #     text = listener.recognize_google(audio, language='en-in')
        #
        #     # with open(user + '.txt', 'a') as file:
        #     #     print((text + '.'), file=file)
        #
        #     print(f'You said :{text}')
        #     friend.say(text)
        #     friend.runAndWait()
        # except:
        #     print('Sorry I couldn\'t recognize your voice')

        # with m as source:
        #     r.adjust_for_ambient_noise(source)
        #     audio = r.listen(source)
        #     value = r.recognize_google(audio, language="en-US")
        #     print(value)





